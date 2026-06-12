import os
import sys
import asyncio
from datetime import datetime

# Add backend directory to path to allow imports from app.*
backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    import subprocess
    print("Installing openpyxl for Excel generation...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from sqlalchemy import text
from app.core.database import async_session_factory

async def fetch_real_order_data():
    """
    Query the database to fetch real orders containing tracking numbers using raw SQL
    to avoid circular import dependency issues in scripts.
    """
    print("🔌 Connecting to database to fetch real tracking data...")
    try:
        async with async_session_factory() as session:
            # Fetch orders with tracking numbers using raw SQL
            stmt = text(
                "SELECT platform, carrier, tracking_number, external_order_id, "
                "fulfillment_channel, ordered_at FROM orders WHERE tracking_number IS NOT NULL;"
            )
            res = await session.execute(stmt)
            orders = res.fetchall()
            print(f"✅ Successfully loaded {len(orders)} orders with tracking numbers from database.")
            return orders
    except Exception as e:
        print(f"⚠️ Could not connect to database ({e}). Falling back to dummy sample data.")
        return []

def populate_sheet1(ws1, orders, thin_border, header_font, header_fill):
    # Tab 1: Daily Pickup Report
    ws1.title = "Daily Pickup Report"
    
    # Title Block
    ws1.merge_cells("A1:I1")
    title_cell = ws1["A1"]
    title_cell.value = "USAV DAILY PICKUP REPORT (BÁO CÁO PICKUP HÀNG NGÀY)"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40
    
    headers1 = [
        "Date (Ngày)", 
        "Carrier (Đơn vị vận chuyển)", 
        "Total Boxes Picked Up (Tổng số thùng đếm thực tế)", 
        "Total Scanned Tracking Numbers (Tổng tracking scan hệ thống)", 
        "Customer Orders Count (Đơn khách lẻ)", 
        "FBA Orders Count (Đơn FBA)", 
        "Checked By (Người đếm)", 
        "Confirmation Photo / Signature (Ảnh xác nhận)",
        "Notes (Ghi chú)"
    ]
    
    for col_idx, header in enumerate(headers1, 1):
        cell = ws1.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws1.row_dimensions[2].height = 28

    # Process and group real data if available, otherwise use fallback
    if orders:
        groups = {}
        for order in orders:
            # order is a Row tuple: (platform, carrier, tracking_number, external_order_id, fulfillment_channel, ordered_at)
            ordered_at = order[5]
            carrier = order[1]
            fulfillment_channel = order[4]

            date_str = ordered_at.strftime("%Y-%m-%d") if ordered_at else datetime.now().strftime("%Y-%m-%d")
            carrier_str = carrier or "USPS"
            is_fba = fulfillment_channel == "AMAZON_FBA"
            
            key = (date_str, carrier_str)
            if key not in groups:
                groups[key] = {"customer": 0, "fba": 0}
            
            if is_fba:
                groups[key]["fba"] += 1
            else:
                groups[key]["customer"] += 1

        # Sort groups by date descending
        sorted_keys = sorted(groups.keys(), key=lambda x: x[0], reverse=True)
        data_rows = []
        for key in sorted_keys:
            date_str, carrier_str = key
            customer_count = groups[key]["customer"]
            fba_count = groups[key]["fba"]
            total_scanned = customer_count + fba_count
            data_rows.append([
                date_str,
                carrier_str,
                total_scanned,
                total_scanned,
                customer_count,
                fba_count,
                "System Export",
                "Verified",
                "Automated database export"
            ])
    else:
        # Fallback dummy data
        data_rows = [
            ["2026-06-12", "UPS", 12, 12, 10, 2, "John Doe", "Signature on file", "Picked up on time"],
            ["2026-06-12", "USPS", 8, 7, 7, 0, "John Doe", "Label photo attached", "Mismatch: 1 missing scan!"],
            ["2026-06-12", "FedEx", 3, 3, 2, 1, "Jane Smith", "Signature on file", ""],
        ]

    for row_idx, row_data in enumerate(data_rows, 3):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left" if col_idx in (2, 7, 8, 9) else "center", vertical="center")
            cell.border = thin_border
        ws1.row_dimensions[row_idx].height = 20

    # Auto-fit columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

def populate_sheet2(ws2, orders, thin_border, header_font):
    # Tab 2: Tracking Verification
    ws2.merge_cells("A1:H1")
    title_cell2 = ws2["A1"]
    title_cell2.value = "USAV TRACKING VERIFICATION (KIỂM TRA CHÉO TRACKING)"
    title_cell2.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    title_cell2.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40
    
    headers2 = [
        "Date Verified (Ngày kiểm tra)",
        "Tracking Number (Số Tracking)",
        "Carrier (Vận chuyển)",
        "Order / Shipment ID (Mã Đơn hàng)",
        "Scanned in System? (Có trong HT?)",
        "Physical Box Count Match? (Khớp SL thùng?)",
        "Verified By (Người kiểm tra)",
        "Verification Status / Discrepancy Notes"
    ]
    
    for col_idx, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[2].height = 28

    if orders:
        data_rows = []
        # Sort by ordered_at descending and limit verification table to last 100 entries for readability
        sorted_orders = sorted(orders, key=lambda o: o[5] or datetime.min, reverse=True)
        for order in sorted_orders[:100]:
            ordered_at = order[5]
            tracking_number = order[2]
            carrier = order[1]
            external_order_id = order[3]
            
            date_str = ordered_at.strftime("%Y-%m-%d") if ordered_at else datetime.now().strftime("%Y-%m-%d")
            data_rows.append([
                date_str,
                tracking_number,
                carrier or "USPS",
                external_order_id,
                "YES",
                "YES",
                "System",
                "Imported"
            ])
    else:
        # Fallback dummy data
        data_rows = [
            ["2026-06-12", "1Z999AA10123456784", "UPS", "13-14737-94831", "YES", "YES", "Alice White", "Verified"],
            ["2026-06-12", "420123459205590123456789012345", "USPS", "4785", "YES", "NO", "Alice White", "Alert: Box was physically picked up but tracking number not marked shipped!"],
        ]

    for row_idx, row_data in enumerate(data_rows, 3):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="left" if col_idx in (2, 4, 8) else "center", vertical="center")
            cell.border = thin_border
        ws2.row_dimensions[row_idx].height = 20
        
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

async def create_template():
    # Fetch database orders
    orders = await fetch_real_order_data()

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin', color='B0C4DE'),
        right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'),
        bottom=Side(style='thin', color='B0C4DE')
    )
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True)

    populate_sheet1(wb.active, orders, thin_border, header_font, header_fill)
    populate_sheet2(wb.create_sheet(title="Tracking Verification"), orders, thin_border, header_font)

    # Save the file
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "USAV_Shipment_Tracking_Template.xlsx")
    wb.save(filepath)
    print(f"📢 Template successfully generated at: {filepath}")

if __name__ == "__main__":
    asyncio.run(create_template())
